import numpy as np
from nltk.probability import FreqDist
import matplotlib.pyplot as plt
import math
import csv
# import scipy
import random
from openpyxl import Workbook
from openpyxl.drawing.image import Image

sample = []

with open('Price-Mileage.csv') as csv_file:  # Читаем выборку из файла 0 работы
    spam_reader = csv.reader(csv_file, quotechar='|')
    for row in spam_reader:
        x, y = row[0].split(';')
        if y.isdigit(): sample.append(int(y))

sample.sort()  # Используя встроенную функцию сортировки получаем ранжированный ряд

R = sample[len(sample) - 1] - sample[0]  # Размах
print("R =", R)

k = round(1 + math.log2(len(sample)))  # Число интервалов(Формула Стёрджеса)
print("k = ", k)

h = round(R / k)  # Длина интервала
print("h =", h)

k += 1  # Иначе интервалы не покроют выборку

x0 = sample[0] - h / 2  # Начало первого частичного интервала
if x0 < 0:
    x0 = 0

print("x0 =", x0)

interval = []
variation = []
x = x0

# Получаем вариационный ряд
for i in range(len(sample)):
    a = 0
    a += 1
    if i == len(sample) - 1:
        variation.append([sample[i], a, a/len(sample)])
        break
    if sample[i] != sample[i+1]:
        variation.append([sample[i], a, a/len(sample)])

a = 0

for i in range(k):
    interval.append([(x, x + h), 0, 0])
    x += h

# Получаем интервальный ряд
for i in sample:
    for j in range(k):
        if interval[j][0][0] < i <= interval[j][0][1]:
            interval[j][1] += 1
            break

for i in interval:
    i[2] = i[1] / len(sample)

print(interval)

middle_int = []
accum_freq = []
accum_afreq = []
accum_freq.append(0)
accum_afreq.append(0)
a = 0
b = 0
c = 0

# Вычисляем серидины интервалов и их накопленные частоты
for i in range(len(interval)):
    a = a + interval[i][2]
    b = b + interval[i][1]
    c = i
    middle_int.append(interval[i][0][0] + h / 2)
    accum_afreq.append(a)
    accum_freq.append(b)

x = middle_int
y = [interval[i][1] for i in range(len(interval))]
plt.plot(x, y)
plt.title("Полигон частот")
plt.savefig('poly.png')
plt.show()
plt.clf()

y = [interval[i][2] for i in range(len(interval))]
plt.plot(x, y)
plt.title("Полигон относительных частот")
plt.savefig("poly_otn.png")
plt.show()
plt.clf()

x = [x0 + i * h + h / 2 for i in range(k)]
y = [interval[i][1] for i in range(len(interval))]
plt.bar(x, y, width=h)
plt.title("Гистограмма частот")
plt.savefig("gist.png")
plt.show()
plt.clf()

y = [interval[i][2] for i in range(len(interval))]
plt.bar(x, y, width=h)
plt.title("Гистограмма относительных частот")
plt.savefig("gist_otn.png")
plt.show()
plt.clf()

X = []
Y = accum_afreq
U = []
V = accum_afreq
for i in range(k + 1):
    U.append(x0 + h * i)
    X.append(x0 + h * i + h)
plt.quiver(X, Y, -h, 0, angles='xy', scale_units='xy', scale=1, color='b')
plt.title("Эмпирическая функция распределения относительных частот")
plt.xlim(0, x0 + h * k)
plt.ylim(0, 1.1)
plt.savefig("emp_func_otn.png")
plt.show()
plt.clf()

Y = accum_freq
V = accum_freq
plt.quiver(X, Y, -h, 0, angles='xy', scale_units='xy', scale=1, color='b')
plt.title("Эмпирическая функция распределения частот")
plt.xlim(0, x0 + h * k)
plt.ylim(0, len(sample) + 10)
plt.savefig("emp_func.png")
plt.show()
plt.clf()


plt.plot(U, accum_freq)
plt.title("Кумулята частот")
plt.savefig("cum.png")
plt.show()
plt.clf()
plt.plot(U, accum_afreq)
plt.title("Кумулята относительных частот")
plt.savefig("cum_otn.png")
plt.show()
plt.clf()

wb = Workbook()
filename = "output.xlsx"
ws = wb.active
ws.title = "Result1"
ws["A1"] = "Ранжированный"
ws.append(sample)  # Делаю так исключительно ради удобства, можно запариться еще больше
ws["A4"] = "Вариационный ряд"
ws.append(variation[i][0] for i in range(len(variation)))
ws["A6"] = "Частоты"
ws.append([variation[i][1] for i in range(len(variation))])
ws["A8"] = "Относительные частоты"
ws.append([variation[i][2] for i in range(len(variation))])

ws["A12"] = "Интервальный ряд"
ws.append([str(interval[i][0]) for i in range(len(interval))])
ws["A14"] = "Частоты"
ws.append([interval[i][1] for i in range(len(interval))])
ws["A16"] = "Относительные частоты"
ws.append([interval[i][2] for i in range(len(interval))])
ws.add_image(Image("poly.png"), "A18")
ws.add_image(Image("poly_otn.png"), "K18")
ws.add_image(Image("gist.png"), "A43")
ws.add_image(Image("gist_otn.png"), "K43")
ws.add_image(Image("emp_func.png"), "A69")
ws.add_image(Image("emp_func_otn.png"), "K69")
ws.add_image(Image("cum.png"), "A95")
ws.add_image(Image("cum_otn.png"), "K95")
wb.save(filename=filename)
