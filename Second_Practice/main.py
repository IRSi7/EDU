import numpy as np
from nltk.probability import FreqDist
import matplotlib.pyplot as plt
import math
import csv
# import scipy
import random
from openpyxl import Workbook
# from openpyxl.drawing.image import Image

sample = []

with open('Price-Mileage.csv') as csv_file:  # Читаем выборку из файла 0 работы
    spam_reader = csv.reader(csv_file, quotechar='|')
    for row in spam_reader:
        x, y = row[0].split(';')
        if y.isdigit(): sample.append(int(y))

sample.sort()  # Используя встроенную функцию сортировки получаем ранжированный ряд

R = sample[len(sample) - 1] - sample[0]  # Размах
print("R =", R)

k = round(1 + math.log2(len(sample)))  # Число интервалов(Формула Стёрджеса)
print("k = ", k)

h = round(R / k)  # Длина интервала
print("h =", h)

k += 1  # Иначе интервалы не покроют выборку

x0 = sample[0] - h / 2  # Начало первого частичного интервала
if x0 < 0:
    x0 = 0

print("x0 =", x0)

interval = []
x = x0

for i in range(k):
    interval.append([(x, x + h), 0, 0])
    x += h

# Получаем интервальный ряд
for i in sample:
    for j in range(k):
        if interval[j][0][0] < i <= interval[j][0][1]:
            interval[j][1] += 1
            break

for i in interval:
    i[2] = i[1] / len(sample)

print(interval)

middle_int = []
accum_freq = []
accum_afreq = []
accum_freq.append(0)
accum_afreq.append(0)
a = 0
b = 0

# Вычисляем серидины интервалов и их накопленные частоты
for i in range(len(interval)):
    a = a + interval[i][2]
    b = b + interval[i][1]
    middle_int.append(interval[i][0][0] + h / 2)
    accum_afreq.append(a)
    accum_freq.append(b)

C = middle_int[int(k/2)]  # Число С из теории, h было вычислено ранее при построении интервального ряда
con_var = []  # Условные варианты

for i in range(k):
    con_var.append([int((middle_int[i] - C)/h), interval[i][1], interval[i][2]])

SEM = []  # Выборочные начальные моменты до 4 порядка
CEM = []  # Выборочные центральные моменты до 4 порядка

# Вычисляем условные эмп момент 1 порядка
for i in range(4):
    SEM.append(0)
    for j in range(k):
        SEM[i] += (con_var[j][0]**(i+1))*con_var[j][2]

CEM.append(0)   # Центральный момент 1 порядка (выборочное среднее для усл вариант)
CEM.append(SEM[1]-SEM[0]**2)  # 2 порядка (выборочная дисперсия для усл вариант)
CEM.append(SEM[2]-3*SEM[1]*SEM[0]+2*SEM[0]**3)   # 3 порядка
CEM.append(SEM[3]-4*SEM[0]*SEM[3]+6*(SEM[0]**2)*SEM[3]-3*SEM[0]**4)  # 4 порядка

Xs = 0      # Выборочное среднее по обычной формуле
CXs = 0     # Выборочное среднее через моменты условных вариант
Ds = 0      # Дисперсия по обычной формуле
CXs = 0     # Дисперсия через моменты условных вариант

for i in range(k):
    Xs += middle_int[i]*interval[i][2]

for i in range(k):
    Ds += (middle_int[i] - Xs)**2*interval[i][2]

CXs = SEM[0]*h + C          # Из формулы метода упрощённых вычислений
CDs = CEM[1]*(h**2)         # Из формулы метода упрощённых вычислений
ds = math.sqrt(Ds)          # Выборочное СКО
cds = math.sqrt(CEM[1])     # Выборочное СКО для усл вариант

Asym = CEM[2]/(cds**3)       # Коэффициэнт ассиметрии для условных вариант (как я понимаю это и есть стат оценка?)
Excess = CEM[3]/(cds**4)     # Коэффициент эксцесса для условных вариант

sDs = Ds*len(sample)/(len(sample)-1)    # Исправленная выборочная дисперсия
s = math.sqrt(sDs)                      # Исправленное выборочное СКО

Mod = 0      # Частость модального интервала
i0 = 0      # Его номер
Mod0 = 0     # Истинное значение моды

for i in range(k):
    if interval[i][1] > Mod:
        Mod = interval[i][1]
        i0 = i

Mod0 = int(interval[i0][0][0] + h*((interval[i0][2] - interval[i0-1][2])/((interval[i0][2] - interval[i0-1][2])+(interval[i0][2] - interval[i0+1][2]))))

i0 = 0      # Номер медианного интервала
Med0 = 0     # Истинное значение медианы

for i in range(k):
    if accum_afreq[i] > 0.5:
        i0 = i
        break

px = 0      # Число нужное для линейной интерполяции медианы
for i in range(i0-1):
    px += 1/interval[i][2]

# Med0 = interval[i0][0][1] + (h/interval[i0][2])*(0.5 - h*px)
# При выполнении линейной интерполяции для медианы получилось неадекватно большое отрицательное число
# Я не знаю с чем это связано поэтому найду её просто как серидину ранжированной ряда

Med0 = sample[int(len(sample)/2)]

CV = Xs/ds      # Коэффициент вариации

wb = Workbook()
ws = wb.create_sheet()

# filename = "output1.xlsx"
# ws = wb.active
ws.title = "Result"
ws["A1"] = "Интервальный"
ws.append([str(interval[i][0]) for i in range(len(interval))])
ws["A3"] = "Середины"
ws.append(middle_int)
ws["A5"] = "Накопленный частоты"
ws.append(accum_afreq[i] for i in range(k))
ws["A7"] = "Условные"
ws.append(con_var[i][0] for i in range(len(interval)))
ws["A9"] = "Начальные эмп моменты"
i = [1, 2, 3, 4]
ws.append(i)
ws.append(SEM)
ws["A12"] = "Центральные эмп моменты"
ws.append(CEM)

ws["A14"] = "Xв через стандартную формулу"
ws["A15"] = Xs
ws["B14"] = "Xв через условные варианты"
ws["B15"] = CXs

ws["A16"] = "Дисперсия через стандартную формулу"
ws["A17"] = Ds
ws["B16"] = "Дисперсия через условные варианты"
ws["B17"] = CDs

ws["A18"] = "Исправленная выборочная дисперсия"
ws["A19"] = sDs
ws["B18"] = "Исправленное выборочное СКО"
ws["B19"] = s
ws["A20"] = "Смещённая оценка дисперсии"
ws["A21"] = Ds
ws["B20"] = "Смещённая оценка СКО"
ws["B21"] = ds

ws["A22"] = "Оценка ассиметрии"
ws["A23"] = Asym
ws["B22"] = "Оценка эксцесса"
ws["B23"] = Excess

ws["A24"] = "Мода"
ws["A25"] = Mod0
ws["B24"] = "Медиана"
ws["B25"] = Med0

ws["A26"] = "Коэффициэнт вариации"
ws["A27"] = CV

wb.save('Result.xlsx')

# TY1 = 1.981      # t гамма для 95% n = 115
# TY2 = 2.62       # t гамма для 99% n = 115
# D1 = TY1*S/math.sqrt(9)
# D2 = TY2*S/math.sqrt(9)
# LB = XV - D1
# LR = XV + D1
# LB2 = XV - D2
# LB2 = XV + D2
# Q1 = D1/S
# Q2 = D2/S
#
# SLB1 = S - S*Q1
# SRB1 = S + S*Q1
# SLB2 = 0
# SRB2 = S + S*Q2
#
# s = []
# for i in range(8):
#     s.append((interval[i][0][1]-XV)/S)
#
# st = [-0.5, -0.2823, -0.0438, 0.2157, 0.3944, 0.4726, 0.4953, 0.49931, 0.499968, 0.5]
# p = []
# for i in range(9):
#     p.append(st[i+1]-st[i])
#
# m = []
# for i in range(9):
#     m.append(p[i]*115)
#
# m2 =[]
# for i in range(9):
#     m2.append(((interval[i][1])**2)/m[i])
#
# ans = sum(m2) - 115
#
# f = []
